from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries, get_column_letter
import json
import os
import shutil
import re
from openai import OpenAI

# Copy the template file first to preserve all formatting
shutil.copy("sample.xlsx", "sample_output.xlsx")

# Now open the copy and work with it
wb = load_workbook("sample_output.xlsx")
ws = wb.active

merged_map = {}

for r in ws.merged_cells.ranges:
    merged_map[r.coord] = r.coord.split(":")[0]  # top-left cell

print("Merged cells found:")
print(f"Total merged ranges: {len(merged_map)}")
for merged_range, top_left in merged_map.items():
    print(f"  {merged_range} -> {top_left}")

if not merged_map:
    print("No merged cells found in the worksheet.")


# Spanish form field keywords commonly found in bank documents
LABEL_KEYWORDS = [
    'nombre', 'dirección', 'teléfono', 'telefono', 'email', 'correo', 'fecha',
    'código', 'codigo', 'actividad', 'representante', 'razón', 'razon', 'social',
    'nit', 'rfc', 'curp', 'rfc', 'ciudad', 'estado', 'país', 'pais', 'cp', 'código postal',
    'banco', 'cuenta', 'clabe', 'swift', 'iban', 'moneda', 'monto', 'importe',
    'apellido', 'paterno', 'materno', 'nacimiento', 'edad', 'género', 'genero',
    'ocupación', 'ocupacion', 'profesión', 'profesion', 'empresa', 'puesto',
    'documento', 'identificación', 'identificacion', 'pasaporte', 'licencia',
    'contacto', 'emergencia', 'parentesco', 'beneficiario', 'titular',
    'firma', 'fecha de', 'lugar de', 'hora', 'folio', 'referencia', 'número', 'numero'
]

# Helper function to check if a cell is part of a merged range (but not the top-left)
def is_merged_cell_but_not_top_left(cell):
    """Check if cell is a MergedCell (not the writable top-left)"""
    return isinstance(cell, MergedCell)

# Helper function to check if a cell is the top-left of a merged range
def is_top_left_of_merged(cell_coord, merged_map):
    """Check if a cell coordinate is the top-left of any merged range"""
    for merged_range, top_left in merged_map.items():
        if cell_coord == top_left:
            return True
    return False

# Helper function to get the cell value safely
def get_cell_value(cell):
    """Safely get cell value, handling MergedCell"""
    if isinstance(cell, MergedCell):
        return None
    return cell.value

# Helper function to check if a cell is empty
def is_cell_empty(cell):
    """Check if a cell is empty or contains only whitespace"""
    if isinstance(cell, MergedCell):
        return True  # Merged cells (non-top-left) are considered empty
    value = cell.value
    return value is None or (isinstance(value, str) and value.strip() == "")

# Helper function to find the field cell in multiple directions
def find_field_cell(label_cell, merged_map):
    """Find the field cell that's associated with a label cell.
    Checks multiple directions: right (most common), below, and left."""
    row_idx = label_cell.row
    col_idx = label_cell.column
    
    # Skip if this cell is a MergedCell (not top-left)
    if is_merged_cell_but_not_top_left(label_cell):
        return None
    
    # Skip if this cell is the top-left of a merged range (it's likely a field value, not a label)
    if is_top_left_of_merged(label_cell.coordinate, merged_map):
        return None
    
    # Try different directions in order of likelihood
    directions = [
        (0, 1),   # Right (most common)
        (1, 0),   # Below
        (0, -1),  # Left (less common, but sometimes labels are on the right)
    ]
    
    for row_offset, col_offset in directions:
        field_row = row_idx + row_offset
        field_col = col_idx + col_offset
        
        # Skip if out of bounds
        if field_row < 1 or field_col < 1:
            continue
        
        # Check if field cell is part of a merged range
        for merged_range, top_left in merged_map.items():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(merged_range)
                if min_row <= field_row <= max_row and min_col <= field_col <= max_col:
                    # Field cell is in this merged range
                    # Check if the top-left of this merged range is empty
                    top_left_cell = ws[top_left]
                    if is_cell_empty(top_left_cell):
                        return merged_range
                    else:
                        break  # This merged range has data, try next direction
            except:
                continue
        
        # Not merged, check single cell
        try:
            field_cell = ws.cell(row=field_row, column=field_col)
            if is_cell_empty(field_cell):
                return field_cell.coordinate
        except:
            pass
    
    return None

# Helper function to determine if text looks like a label (not a field value)
def looks_like_label(text):
    """Determine if text looks like a form label rather than a field value"""
    if not text or not isinstance(text, str):
        return False
    
    text = text.strip()
    
    # Too short
    if len(text) < 2:
        return False
    
    text_lower = text.lower()
    
    # Must end with colon OR contain form keywords
    ends_with_colon = text.rstrip().endswith(':')
    has_keyword = any(keyword in text_lower for keyword in LABEL_KEYWORDS)
    
    if not (ends_with_colon or has_keyword):
        return False
    
    # If it ends with colon, it's very likely a label (but check a few edge cases)
    if ends_with_colon:
        # Very long labels with colons are still labels, but check for obvious field values
        if len(text) > 100:
            return False
        # If it looks like a URL or email with colon, skip
        if '@' in text and not any(kw in text_lower for kw in ['email', 'correo', 'mail']):
            return False
        if text_lower.startswith(('http', 'www')):
            return False
        return True
    
    # For labels without colons, apply stricter rules
    # Labels are usually shorter
    if len(text) > 40:
        return False
    
    # Additional heuristics: labels are usually shorter and don't contain certain patterns
    # Field values often contain:
    # - Numbers in the middle (like addresses, phone numbers)
    # - Special characters like @, /, -, (
    # - Very long text
    
    # If it has numbers in the middle (not just at start/end), it's likely a field value
    if re.search(r'\d+', text) and not text[0].isdigit() and not text[-1].isdigit():
        return False
    
    # If it contains @, it's likely an email field value, not a label
    if '@' in text and 'email' not in text_lower and 'correo' not in text_lower:
        return False
    
    # If it starts with "http" or "www", it's a URL field value
    if text_lower.startswith(('http', 'www')):
        return False
    
    # If it contains common field value patterns (like phone numbers, addresses)
    if re.search(r'\d{3,}', text):  # Long sequences of numbers
        return False
    
    return True

# Collect all potential fields with improved detection
all_potential_fields = {}
seen_coordinates = set()  # Track which field cells we've already mapped

print("\n=== SCANNING FOR FORM FIELDS ===")

for row in ws.iter_rows():
    for cell in row:
        if not cell.value or not isinstance(cell.value, str):
            continue
        
        # Skip if this cell is a MergedCell (not top-left)
        if is_merged_cell_but_not_top_left(cell):
            continue
        
        # Skip if this cell is the top-left of a merged range (likely a field value)
        if is_top_left_of_merged(cell.coordinate, merged_map):
            continue
        
        label = cell.value.strip()
        
        # Check if this looks like a label
        if not looks_like_label(label):
            continue
        
        row_idx = cell.row
        col_idx = cell.column
        
        # Additional check: If there's significant text to the left, this might be a field value
        # Labels are usually in the leftmost columns or have empty cells to their left
        if col_idx > 1:  # Not in first column
            try:
                left_cell = ws.cell(row=row_idx, column=col_idx - 1)
                left_value = get_cell_value(left_cell)
                # If left cell has substantial text, current cell might be a field value
                if left_value and isinstance(left_value, str) and len(left_value.strip()) > 10:
                    # But if current cell ends with colon, it's still likely a label
                    if not label.rstrip().endswith(':'):
                        continue
            except:
                pass
        
        # Find the field cell (checks multiple directions)
        target = find_field_cell(cell, merged_map)
        
        if target is None:
            continue  # No valid field cell found
        
        # Get the top-left coordinate of the target (for deduplication)
        if ":" in target:
            target_top_left = target.split(":")[0]
        else:
            target_top_left = target
        
        # Verify the field cell is actually empty
        try:
            if ":" in target:
                top_left = target.split(":")[0]
                field_cell = ws[top_left]
            else:
                field_cell = ws[target]
            
            if not is_cell_empty(field_cell):
                continue  # Field already has data
        except Exception as e:
            continue  # Skip if we can't verify
        
        # Check for duplicate field cells
        if target_top_left in seen_coordinates:
            # This field cell is already mapped
            # Only replace if the new label is more specific (ends with colon)
            existing_label = None
            for lbl, tgt in all_potential_fields.items():
                tgt_tl = tgt.split(":")[0] if ":" in tgt else tgt
                if tgt_tl == target_top_left:
                    existing_label = lbl
                    break
            
            if existing_label:
                # If new label ends with colon and old one doesn't, replace
                if label.rstrip().endswith(':') and not existing_label.rstrip().endswith(':'):
                    del all_potential_fields[existing_label]
                else:
                    continue  # Keep existing mapping
        
        # All checks passed - this is a valid label -> field mapping
        # Use the first occurrence if duplicate labels exist
        if label not in all_potential_fields:
            all_potential_fields[label] = target
            seen_coordinates.add(target_top_left)

print("\n=== FIELD DETECTION ANALYSIS ===")
print(f"Total potential fields found (after improved filtering): {len(all_potential_fields)}")

# Final fields dictionary (already filtered)
fields = all_potential_fields

print(f"Total fields found: {len(fields)}")

print("\nDetected fields (label -> field cell):")
for label, target in fields.items():
    print(f"  '{label}' -> {target}")

if not fields:
    print("No fields found in the worksheet after filtering.")


def get_data_from_ai(field_labels, max_retries=3):
    """
    Calls AI API to generate JSON data based on the field labels.
    Returns a dictionary with field labels as keys and generated values.
    """
    # Initialize OpenAI client (requires OPENAI_API_KEY environment variable)
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable is not set. Please set it to use AI data generation.")
    
    client = OpenAI(api_key=api_key)
    
    # Create prompt for AI - be more explicit about JSON format
    field_list = ", ".join(field_labels)
    prompt = f"""Generate realistic sample data for the following form fields: {field_list}

IMPORTANT: Return ONLY a complete, valid JSON object. The JSON must:
1. Have each field name as a key (exactly as listed above)
2. Have appropriate sample data as the value for each key
3. Be complete and properly closed (all brackets and braces must be closed)
4. Be valid JSON that can be parsed

Return ONLY the JSON object, no explanations, no markdown, no code blocks.

Example format:
{{"Full Name": "John Doe", "Address": "123 Main St, City, Country", "DOB": "01-15-1990", "Gender": "Male"}}
"""
    
    print("\nCalling AI to generate data...")
    print(f"Fields sent to AI: {field_list}")
    print(f"Total fields: {len(field_labels)}")
    
    for attempt in range(max_retries):
        try:
            # Calculate appropriate max_tokens based on number of fields
            # Estimate: ~50-100 tokens per field + buffer for structure
            estimated_tokens = max(2000, len(field_labels) * 80 + 1000)
            print(f"Using max_tokens: {estimated_tokens} (attempt {attempt + 1}/{max_retries})")
            
            # Try with JSON mode first, fallback if not supported
            try:
                response = client.chat.completions.create(
                    model="gpt-4.1-mini",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that generates realistic sample data. Always return complete, valid JSON objects. Never return incomplete or truncated JSON. Ensure all field names from the user's list are included as keys."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2,
                    max_tokens=estimated_tokens,
                    response_format={"type": "json_object"}  # Force JSON mode
                )
            except Exception as json_mode_error:
                # If JSON mode not supported, try without it
                print(f"JSON mode not supported, trying without it: {json_mode_error}")
                response = client.chat.completions.create(
                    model="gpt-4.1-mini",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that generates realistic sample data. Always return complete, valid JSON objects. Never return incomplete or truncated JSON. Ensure all field names from the user's list are included as keys."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2,
                    max_tokens=estimated_tokens
                )
            
            # Extract JSON from response
            ai_response = response.choices[0].message.content.strip()
            
            # Remove markdown code blocks if present
            if ai_response.startswith("```"):
                parts = ai_response.split("```")
                # Find the JSON part (usually the middle part)
                for part in parts:
                    part = part.strip()
                    if part.startswith("json"):
                        part = part[4:].strip()
                    if part.startswith("{") or part.startswith("["):
                        ai_response = part
                        break
                else:
                    # If no JSON found, try the first non-empty part
                    ai_response = next((p.strip() for p in parts if p.strip() and not p.strip().startswith("json")), ai_response)
            
            # Try to fix incomplete JSON if needed
            ai_response = ai_response.strip()
            if not ai_response.endswith("}") and not ai_response.endswith("]"):
                # Try to find and close incomplete JSON
                open_braces = ai_response.count("{") - ai_response.count("}")
                open_brackets = ai_response.count("[") - ai_response.count("]")
                
                if open_braces > 0 or open_brackets > 0:
                    print(f"Warning: JSON appears incomplete. Attempting to fix...")
                    # Close any open structures
                    ai_response += "}" * open_braces + "]" * open_brackets
            
            # Parse JSON
            data = json.loads(ai_response)
            
            print(f"AI Response received successfully: {len(data)} fields")
            print(f"AI Response preview (first 3 fields): {json.dumps(dict(list(data.items())[:3]), indent=2)}...")
            return data
            
        except json.JSONDecodeError as e:
            print(f"Error parsing AI response as JSON (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                print(f"Retrying with increased token limit...")
                # Increase tokens on retry
                estimated_tokens = int(estimated_tokens * 1.5)
                continue
            else:
                print(f"Raw AI response (last 500 chars): {ai_response[-500:]}")
                raise ValueError(f"Failed to parse AI response as JSON after {max_retries} attempts. The response may be incomplete. Last error: {e}")
        except Exception as e:
            print(f"Error calling AI API (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                print(f"Retrying...")
                continue
            raise


# Get field labels (keys from fields dictionary)
field_labels = list(fields.keys())

if field_labels:
    # Call AI to get dynamic data
    data = get_data_from_ai(field_labels)
    
    print("\nFilling Excel with AI-generated data...")
    
    # Helper function to check if a cell coordinate is within a merged range
    def is_cell_in_range(cell_coord, merged_range):
        """Check if a cell coordinate is within a merged range"""
        if ":" not in merged_range:
            return cell_coord == merged_range
        
        start, end = merged_range.split(":")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(merged_range)
            cell_col = ws[cell_coord].column
            cell_row = ws[cell_coord].row
            return min_col <= cell_col <= max_col and min_row <= cell_row <= max_row
        except:
            # Fallback: simple string check
            return cell_coord == start or (start <= cell_coord <= end)
    
    # Helper function to get the writable cell for a target (handles merged cells)
    def get_writable_cell(target_coord):
        """Get the actual writable cell for a target coordinate (handles merged cells)"""
        if ":" in target_coord:
            # This is a merged range
            top_left = target_coord.split(":")[0]
            # Check if it's in merged_map
            if target_coord in merged_map:
                return ws[merged_map[target_coord]]
            else:
                # Try to find it in merged_cells
                for merged_range_obj in ws.merged_cells.ranges:
                    if merged_range_obj.coord == target_coord:
                        return ws.cell(row=merged_range_obj.min_row, column=merged_range_obj.min_col)
                # Fallback to top-left
                return ws[top_left]
        else:
            # Single cell
            cell = ws[target_coord]
            # If it's a MergedCell, find the top-left
            if isinstance(cell, MergedCell):
                for merged_range, top_left in merged_map.items():
                    if is_cell_in_range(target_coord, merged_range):
                        return ws[top_left]
                # If we can't find it, try merged_cells directly
                for merged_range_obj in ws.merged_cells.ranges:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(merged_range_obj.coord)
                        cell_col = ws[target_coord].column
                        cell_row = ws[target_coord].row
                        if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
                            return ws.cell(row=min_row, column=min_col)
                    except:
                        continue
            return cell
    
    # Fill Excel with the AI-generated data
    filled_count = 0
    skipped_count = 0
    
    for label, target in fields.items():
        try:
            if label not in data:
                print(f"  Warning: '{label}' not found in AI response data")
                skipped_count += 1
                continue
            
            value = data[label]
            
            # Convert complex objects (dict, list) to JSON string
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            # Convert None to empty string
            elif value is None:
                value = ""
            else:
                # Convert to string
                value = str(value)
            
            # Get the writable cell
            try:
                cell = get_writable_cell(target)
            except Exception as e:
                print(f"  Error getting writable cell for '{label}' -> {target}: {e}")
                skipped_count += 1
                continue
            
            # Verify the cell is still empty (double-check)
            if not is_cell_empty(cell):
                print(f"  Skipped '{label}' -> {cell.coordinate} (cell already has data: '{str(cell.value)[:30]}...')")
                skipped_count += 1
                continue
            
            # Set the value on the writable cell
            cell.value = value
            filled_count += 1
            print(f"  ✓ Filled '{label}' -> {cell.coordinate} ({target}) with: {str(value)[:50]}...")
            
        except Exception as e:
            print(f"  ✗ Error filling '{label}': {e}")
            import traceback
            traceback.print_exc()
            skipped_count += 1
            continue  # Continue with next field even if this one fails
    
    print(f"\nSummary: {filled_count} fields filled, {skipped_count} fields skipped")
    
    # Save the output file
    output_filename = "sample_output.xlsx"
    wb.save(output_filename)
    print(f"\nExcel file saved as: {output_filename}")
else:
    print("\nNo fields to process. Skipping AI data generation.")