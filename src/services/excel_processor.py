"""Excel processing service for field detection and manipulation."""
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from typing import Dict, Set, Optional
from src.domain.models import FormField, FieldDetectionResult
from src.domain.exceptions import FieldDetectionError
from src.utils.cell_helpers import (
    is_merged_cell_but_not_top_left,
    is_top_left_of_merged,
    get_cell_value,
    is_cell_empty,
)
from src.utils.label_detector import looks_like_label
from src.config import config


class ExcelProcessor:
    """Service for processing Excel files and detecting form fields."""
    
    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.worksheet = workbook.active
    
    def build_merged_map(self) -> Dict[str, str]:
        """Build a map of merged cell ranges to their top-left coordinates."""
        merged_map = {}
        for merged_range in self.worksheet.merged_cells.ranges:
            merged_map[merged_range.coord] = merged_range.coord.split(":")[0]
        return merged_map
    
    def find_field_cell(self, label_cell, merged_map: Dict[str, str]) -> Optional[str]:
        """Find the field cell associated with a label cell."""
        row_idx = label_cell.row
        col_idx = label_cell.column
        
        if is_merged_cell_but_not_top_left(label_cell):
            return None
        
        if is_top_left_of_merged(label_cell.coordinate, merged_map):
            return None
        
        directions = [(0, 1), (1, 0), (0, -1)]  # Right, Below, Left
        
        for row_offset, col_offset in directions:
            field_row = row_idx + row_offset
            field_col = col_idx + col_offset
            
            if field_row < 1 or field_col < 1:
                continue
            
            for merged_range, top_left in merged_map.items():
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(merged_range)
                    if min_row <= field_row <= max_row and min_col <= field_col <= max_col:
                        top_left_cell = self.worksheet[top_left]
                        if is_cell_empty(top_left_cell):
                            return merged_range
                        else:
                            break
                except Exception:
                    continue
            
            try:
                field_cell = self.worksheet.cell(row=field_row, column=field_col)
                if is_cell_empty(field_cell):
                    return field_cell.coordinate
            except Exception:
                pass
        
        return None
    
    def detect_fields(self) -> FieldDetectionResult:
        """Detect form fields in the Excel workbook."""
        merged_map = self.build_merged_map()
        all_potential_fields: Dict[str, str] = {}
        seen_coordinates: Set[str] = set()
        cfg = config.label_detection
        
        for row in self.worksheet.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                
                if is_merged_cell_but_not_top_left(cell):
                    continue
                
                if is_top_left_of_merged(cell.coordinate, merged_map):
                    continue
                
                label = cell.value.strip()
                
                if not looks_like_label(label):
                    continue
                
                row_idx = cell.row
                col_idx = cell.column
                
                if col_idx > 1:
                    try:
                        left_cell = self.worksheet.cell(row=row_idx, column=col_idx - 1)
                        left_value = get_cell_value(left_cell)
                        if (left_value and isinstance(left_value, str) and 
                            len(left_value.strip()) > cfg.min_left_cell_text_length):
                            if not label.rstrip().endswith(':'):
                                continue
                    except Exception:
                        pass
                
                target = self.find_field_cell(cell, merged_map)
                
                if target is None:
                    continue
                
                target_top_left = target.split(":")[0] if ":" in target else target
                
                try:
                    if ":" in target:
                        top_left = target.split(":")[0]
                        field_cell = self.worksheet[top_left]
                    else:
                        field_cell = self.worksheet[target]
                    
                    if not is_cell_empty(field_cell):
                        continue
                except Exception:
                    continue
                
                if target_top_left in seen_coordinates:
                    existing_label = None
                    for lbl, tgt in all_potential_fields.items():
                        tgt_tl = tgt.split(":")[0] if ":" in tgt else tgt
                        if tgt_tl == target_top_left:
                            existing_label = lbl
                            break
                    
                    if existing_label:
                        if (label.rstrip().endswith(':') and 
                            not existing_label.rstrip().endswith(':')):
                            del all_potential_fields[existing_label]
                        else:
                            continue
                
                if label not in all_potential_fields:
                    all_potential_fields[label] = target
                    seen_coordinates.add(target_top_left)
        
        fields = {
            label: FormField(label=label, target_coordinate=target)
            for label, target in all_potential_fields.items()
        }
        
        return FieldDetectionResult(fields=fields, merged_map=merged_map)

