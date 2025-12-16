"""Utility functions for Excel cell operations."""
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional


def is_merged_cell_but_not_top_left(cell: Cell) -> bool:
    """Check if cell is a MergedCell (not the writable top-left)."""
    return isinstance(cell, MergedCell)


def is_top_left_of_merged(cell_coord: str, merged_map: dict[str, str]) -> bool:
    """Check if a cell coordinate is the top-left of any merged range."""
    return cell_coord in merged_map.values()


def get_cell_value(cell: Cell) -> Optional[str]:
    """Safely get cell value, handling MergedCell."""
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def is_cell_empty(cell: Cell) -> bool:
    """Check if a cell is empty or contains only whitespace."""
    if isinstance(cell, MergedCell):
        return True
    value = cell.value
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_cell_in_range(cell_coord: str, merged_range: str, ws: Worksheet) -> bool:
    """Check if a cell coordinate is within a merged range."""
    if ":" not in merged_range:
        return cell_coord == merged_range
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(merged_range)
        cell = ws[cell_coord]
        return min_col <= cell.column <= max_col and min_row <= cell.row <= max_row
    except Exception:
        start = merged_range.split(":")[0]
        return cell_coord == start


def get_writable_cell(target_coord: str, ws: Worksheet, merged_map: dict[str, str]) -> Cell:
    """Get the actual writable cell for a target coordinate (handles merged cells)."""
    if ":" in target_coord:
        top_left = target_coord.split(":")[0]
        if target_coord in merged_map:
            return ws[merged_map[target_coord]]
        else:
            for merged_range_obj in ws.merged_cells.ranges:
                if merged_range_obj.coord == target_coord:
                    return ws.cell(row=merged_range_obj.min_row, column=merged_range_obj.min_col)
            return ws[top_left]
    else:
        cell = ws[target_coord]
        if isinstance(cell, MergedCell):
            for merged_range, top_left in merged_map.items():
                if is_cell_in_range(target_coord, merged_range, ws):
                    return ws[top_left]
            for merged_range_obj in ws.merged_cells.ranges:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(merged_range_obj.coord)
                    cell_col = cell.column
                    cell_row = cell.row
                    if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
                        return ws.cell(row=min_row, column=min_col)
                except Exception:
                    continue
        return cell

