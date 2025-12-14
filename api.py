from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
from fastapi.middleware.cors import CORSMiddleware
import json
import os
import re
from openai import OpenAI
from io import BytesIO
from typing import Dict, List, Optional
import uuid
import tempfile

app = FastAPI(title="Excel Form Filler API", description="API to detect and fill form fields in Excel files using AI")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with specific origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Spanish form field keywords commonly found in bank documents
LABEL_KEYWORDS = [
    'nombre', 'dirección', 'teléfono', 'telefono', 'email', 'correo', 'fecha',
    'código', 'codigo', 'actividad', 'representante', 'razón', 'razon', 'social',
    'nit', 'rfc', 'curp', 'rfc', 'ciudad', 'estado', 'país', 'pais', 'cp', 'código postal',
    'banco', 'cuenta', 'clabe', 'swift', 'iban', 'moneda', 'monto', 'importe',
    'apellido', 'paterno', 'materno', 'nacimiento', 'edad', 'género', 'genero',
    'ocupación', 'ocupacion', 'profesión', 'profesion', 'empresa', 'puesto',
    'documento', 'identificación', 'identificacion', 'pasaporte', 'licencia',
    'contacto', 'emergencia', 'parentesco', 'beneficiario', 'titular',
    'firma', 'fecha de', 'lugar de', 'hora', 'folio', 'referencia', 'número', 'numero'
]

# Helper functions
def is_merged_cell_but_not_top_left(cell):
    """Check if cell is a MergedCell (not the writable top-left)"""
    return isinstance(cell, MergedCell)

def is_top_left_of_merged(cell_coord, merged_map):
    """Check if a cell coordinate is the top-left of any merged range"""
    for merged_range, top_left in merged_map.items():
        if cell_coord == top_left:
            return True
    return False

def get_cell_value(cell):
    """Safely get cell value, handling MergedCell"""
    if isinstance(cell, MergedCell):
        return None
    return cell.value

def is_cell_empty(cell):
    """Check if a cell is empty or contains only whitespace"""
    if isinstance(cell, MergedCell):
        return True
    value = cell.value
    return value is None or (isinstance(value, str) and value.strip() == "")

def find_field_cell(label_cell, merged_map, ws):
    """Find the field cell that's associated with a label cell."""
    row_idx = label_cell.row
    col_idx = label_cell.column
    
    if is_merged_cell_but_not_top_left(label_cell):
        return None
    
    if is_top_left_of_merged(label_cell.coordinate, merged_map):
        return None
    
    directions = [(0, 1), (1, 0), (0, -1)]  # Right, Below, Left
    
    for row_offset, col_offset in directions:
        field_row = row_idx + row_offset
        field_col = col_idx + col_offset
        
        if field_row < 1 or field_col < 1:
            continue
        
        for merged_range, top_left in merged_map.items():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(merged_range)
                if min_row <= field_row <= max_row and min_col <= field_col <= max_col:
                    top_left_cell = ws[top_left]
                    if is_cell_empty(top_left_cell):
                        return merged_range
                    else:
                        break
            except:
                continue
        
        try:
            field_cell = ws.cell(row=field_row, column=field_col)
            if is_cell_empty(field_cell):
                return field_cell.coordinate
        except:
            pass
    
    return None

def looks_like_label(text):
    """Determine if text looks like a form label rather than a field value"""
    if not text or not isinstance(text, str):
        return False
    
    text = text.strip()
    
    if len(text) < 2:
        return False
    
    text_lower = text.lower()
    ends_with_colon = text.rstrip().endswith(':')
    has_keyword = any(keyword in text_lower for keyword in LABEL_KEYWORDS)
    
    if not (ends_with_colon or has_keyword):
        return False
    
    if ends_with_colon:
        if len(text) > 100:
            return False
        if '@' in text and not any(kw in text_lower for kw in ['email', 'correo', 'mail']):
            return False
        if text_lower.startswith(('http', 'www')):
            return False
        return True
    
    if len(text) > 40:
        return False
    
    if re.search(r'\d+', text) and not text[0].isdigit() and not text[-1].isdigit():
        return False
    
    if '@' in text and 'email' not in text_lower and 'correo' not in text_lower:
        return False
    
    if text_lower.startswith(('http', 'www')):
        return False
    
    if re.search(r'\d{3,}', text):
        return False
    
    return True

def detect_fields(wb):
    """Detect form fields in the Excel workbook"""
    ws = wb.active
    merged_map = {}
    
    for r in ws.merged_cells.ranges:
        merged_map[r.coord] = r.coord.split(":")[0]
    
    all_potential_fields = {}
    seen_coordinates = set()
    
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue
            
            if is_merged_cell_but_not_top_left(cell):
                continue
            
            if is_top_left_of_merged(cell.coordinate, merged_map):
                continue
            
            label = cell.value.strip()
            
            if not looks_like_label(label):
                continue
            
            row_idx = cell.row
            col_idx = cell.column
            
            if col_idx > 1:
                try:
                    left_cell = ws.cell(row=row_idx, column=col_idx - 1)
                    left_value = get_cell_value(left_cell)
                    if left_value and isinstance(left_value, str) and len(left_value.strip()) > 10:
                        if not label.rstrip().endswith(':'):
                            continue
                except:
                    pass
            
            target = find_field_cell(cell, merged_map, ws)
            
            if target is None:
                continue
            
            if ":" in target:
                target_top_left = target.split(":")[0]
            else:
                target_top_left = target
            
            try:
                if ":" in target:
                    top_left = target.split(":")[0]
                    field_cell = ws[top_left]
                else:
                    field_cell = ws[target]
                
                if not is_cell_empty(field_cell):
                    continue
            except Exception as e:
                continue
            
            if target_top_left in seen_coordinates:
                existing_label = None
                for lbl, tgt in all_potential_fields.items():
                    tgt_tl = tgt.split(":")[0] if ":" in tgt else tgt
                    if tgt_tl == target_top_left:
                        existing_label = lbl
                        break
                
                if existing_label:
                    if label.rstrip().endswith(':') and not existing_label.rstrip().endswith(':'):
                        del all_potential_fields[existing_label]
                    else:
                        continue
            
            if label not in all_potential_fields:
                all_potential_fields[label] = target
                seen_coordinates.add(target_top_left)
    
    return all_potential_fields, merged_map

def get_data_from_ai(field_labels, max_retries=3):
    """Calls AI API to generate JSON data based on the field labels."""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY environment variable is not set")
    
    client = OpenAI(api_key=api_key)
    
    field_list = ", ".join(field_labels)
    prompt = f"""Generate realistic sample data for the following form fields: {field_list}

IMPORTANT: Return ONLY a complete, valid JSON object. The JSON must:
1. Have each field name as a key (exactly as listed above)
2. Have appropriate sample data as the value for each key
3. Be complete and properly closed (all brackets and braces must be closed)
4. Be valid JSON that can be parsed

Return ONLY the JSON object, no explanations, no markdown, no code blocks.

Example format:
{{"Full Name": "John Doe", "Address": "123 Main St, City, Country", "DOB": "01-15-1990", "Gender": "Male"}}
"""
    
    for attempt in range(max_retries):
        try:
            estimated_tokens = max(2000, len(field_labels) * 80 + 1000)
            
            try:
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that generates realistic sample data. Always return complete, valid JSON objects. Never return incomplete or truncated JSON. Ensure all field names from the user's list are included as keys."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2,
                    max_tokens=estimated_tokens,
                    response_format={"type": "json_object"}
                )
            except Exception as json_mode_error:
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that generates realistic sample data. Always return complete, valid JSON objects. Never return incomplete or truncated JSON. Ensure all field names from the user's list are included as keys."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2,
                    max_tokens=estimated_tokens
                )
            
            ai_response = response.choices[0].message.content.strip()
            
            if ai_response.startswith("```"):
                parts = ai_response.split("```")
                for part in parts:
                    part = part.strip()
                    if part.startswith("json"):
                        part = part[4:].strip()
                    if part.startswith("{") or part.startswith("["):
                        ai_response = part
                        break
                else:
                    ai_response = next((p.strip() for p in parts if p.strip() and not p.strip().startswith("json")), ai_response)
            
            ai_response = ai_response.strip()
            if not ai_response.endswith("}") and not ai_response.endswith("]"):
                open_braces = ai_response.count("{") - ai_response.count("}")
                open_brackets = ai_response.count("[") - ai_response.count("]")
                
                if open_braces > 0 or open_brackets > 0:
                    ai_response += "}" * open_braces + "]" * open_brackets
            
            data = json.loads(ai_response)
            return data
            
        except json.JSONDecodeError as e:
            if attempt < max_retries - 1:
                estimated_tokens = int(estimated_tokens * 1.5)
                continue
            raise ValueError(f"Failed to parse AI response as JSON after {max_retries} attempts: {e}")
        except Exception as e:
            if attempt < max_retries - 1:
                continue
            raise

def is_cell_in_range(cell_coord, merged_range, ws):
    """Check if a cell coordinate is within a merged range"""
    if ":" not in merged_range:
        return cell_coord == merged_range
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(merged_range)
        cell = ws[cell_coord]
        cell_col = cell.column
        cell_row = cell.row
        return min_col <= cell_col <= max_col and min_row <= cell_row <= max_row
    except:
        start, end = merged_range.split(":")
        return cell_coord == start or (start <= cell_coord <= end)

def get_writable_cell(target_coord, ws, merged_map):
    """Get the actual writable cell for a target coordinate (handles merged cells)"""
    if ":" in target_coord:
        top_left = target_coord.split(":")[0]
        if target_coord in merged_map:
            return ws[merged_map[target_coord]]
        else:
            for merged_range_obj in ws.merged_cells.ranges:
                if merged_range_obj.coord == target_coord:
                    return ws.cell(row=merged_range_obj.min_row, column=merged_range_obj.min_col)
            return ws[top_left]
    else:
        cell = ws[target_coord]
        if isinstance(cell, MergedCell):
            for merged_range, top_left in merged_map.items():
                if is_cell_in_range(target_coord, merged_range, ws):
                    return ws[top_left]
            for merged_range_obj in ws.merged_cells.ranges:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(merged_range_obj.coord)
                    cell_col = cell.column
                    cell_row = cell.row
                    if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
                        return ws.cell(row=min_row, column=min_col)
                except:
                    continue
        return cell

def fill_excel(wb, fields, data, merged_map):
    """Fill Excel workbook with generated data"""
    ws = wb.active
    filled_count = 0
    skipped_count = 0
    errors = []
    
    for label, target in fields.items():
        try:
            if label not in data:
                errors.append(f"'{label}' not found in AI response data")
                skipped_count += 1
                continue
            
            value = data[label]
            
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            elif value is None:
                value = ""
            else:
                value = str(value)
            
            try:
                cell = get_writable_cell(target, ws, merged_map)
            except Exception as e:
                errors.append(f"Error getting writable cell for '{label}' -> {target}: {e}")
                skipped_count += 1
                continue
            
            if not is_cell_empty(cell):
                skipped_count += 1
                continue
            
            cell.value = value
            filled_count += 1
            
        except Exception as e:
            errors.append(f"Error filling '{label}': {e}")
            skipped_count += 1
            continue
    
    return filled_count, skipped_count, errors

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Excel Form Filler API",
        "description": "Upload an Excel file, automatically detect form fields, and fill them with AI-generated data",
        "main_endpoint": {
            "url": "/process",
            "method": "POST",
            "description": "Upload Excel file → Detect fields → Fill with AI data → Return filled file"
        },
        "other_endpoints": {
            "/detect": "POST - Detect form fields in Excel file (returns field list)",
            "/fill": "POST - Fill Excel file with AI-generated or custom data"
        }
    }

@app.post("/process")
async def process_excel(file: UploadFile = File(...)):
    """
    **Main Endpoint** - Complete workflow: Upload Excel → Detect fields → Fill with AI → Return filled file.
    
    This endpoint:
    1. Accepts an Excel file upload
    2. Automatically detects form fields in the file
    3. Generates realistic data using AI for each detected field
    4. Fills the Excel file with the generated data
    5. Returns the filled Excel file as a download
    
    **Parameters:**
    - `file`: Excel file (.xlsx or .xls) to process
    
    **Returns:**
    - Filled Excel file as download (preserves all formatting and merged cells)
    
    **Example:**
    ```bash
    curl -X POST "http://localhost:8000/process" \\
      -F "file=@your_form.xlsx" \\
      -o filled_form.xlsx
    ```
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        # Read uploaded file
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        
        # Step 1: Detect fields
        fields, merged_map = detect_fields(wb)
        
        if not fields:
            raise HTTPException(
                status_code=400, 
                detail="No form fields detected in the file. Make sure your Excel file contains form labels (text ending with ':' or containing form keywords)."
            )
        
        # Step 2: Generate data using AI
        field_labels = list(fields.keys())
        data = get_data_from_ai(field_labels)
        
        # Step 3: Fill Excel with generated data
        filled_count, skipped_count, errors = fill_excel(wb, fields, data, merged_map)
        
        # Step 4: Save and return the filled file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=filled_{file.filename}",
                "X-Fields-Detected": str(len(fields)),
                "X-Fields-Filled": str(filled_count),
                "X-Fields-Skipped": str(skipped_count)
            }
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.post("/detect")
async def detect_form_fields(file: UploadFile = File(...)):
    """Detect form fields in an uploaded Excel file (returns field list without filling)"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        fields, merged_map = detect_fields(wb)
        
        return {
            "filename": file.filename,
            "fields_detected": len(fields),
            "fields": {label: target for label, target in fields.items()}
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.post("/fill")
async def fill_form(
    file: UploadFile = File(...),
    use_ai: bool = True,
    custom_data: Optional[str] = None
):
    """
    Fill Excel file with data (advanced endpoint with custom data option).
    
    - **file**: Excel file to fill
    - **use_ai**: If True, use AI to generate data. If False, use custom_data
    - **custom_data**: JSON string with field labels as keys and values to fill
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        fields, merged_map = detect_fields(wb)
        
        if not fields:
            raise HTTPException(status_code=400, detail="No form fields detected in the file")
        
        # Get data
        if use_ai:
            field_labels = list(fields.keys())
            data = get_data_from_ai(field_labels)
        else:
            if not custom_data:
                raise HTTPException(status_code=400, detail="custom_data required when use_ai is False")
            data = json.loads(custom_data)
        
        # Fill Excel
        filled_count, skipped_count, errors = fill_excel(wb, fields, data, merged_map)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=filled_{file.filename}"}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
